"""
report.py — Generate the final evaluation Excel report matching the required format.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "测评ID",        # A
    "任务名称",      # B
    "任务描述",      # C
    "语言模型",      # D
    "涉及Agent",     # E
    "对话轮次",      # F
    "AI提问消耗Token",  # G
    "AI回答消耗Token",  # H
    "总消耗Token",      # I
    "代码总行数",       # J
    "代码文件数",       # K
    "自我迭代次数",     # L
    "总耗时",           # M
    "运行结果截图",     # N
]

COLUMN_WIDTHS = [8, 14, 28, 16, 22, 10, 18, 18, 14, 12, 12, 14, 10, 16]


def generate_report(results: list[dict], output_path: str):
    """
    results: list of dicts with keys matching HEADERS above.
    output_path: where to save the .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "评测结果"

    # -- Styles --
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    alt_fill = PatternFill("solid", fgColor="D9E2F3")

    # -- Header row --
    # Merge row 1 for top-level grouping: columns G-I get a merged "测评ID" header
    ws.merge_cells("G1:I1")
    ws["G1"] = "测评ID"
    ws["G1"].font = header_font
    ws["G1"].fill = header_fill
    ws["G1"].alignment = header_align

    # Row 2 is the actual header
    for col_idx, (header, width) in enumerate(zip(HEADERS, COLUMN_WIDTHS), 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 1 for non-merged columns just gets the same fill
    for col_idx in range(1, len(HEADERS) + 1):
        if col_idx < 7 or col_idx > 9:
            c = ws.cell(row=1, column=col_idx)
            c.fill = header_fill
            c.border = thin_border

    # -- Data rows --
    for row_idx, result in enumerate(results, 3):
        is_alt = (row_idx % 2 == 1)
        for col_idx, key in enumerate(HEADERS, 1):
            value = result.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill

    # -- Insert screenshots as images --
    try:
        from openpyxl.drawing.image import Image as XlImage
        for row_idx, result in enumerate(results, 3):
            screenshot_path = result.get("_screenshot_path", "")
            if screenshot_path and os.path.exists(screenshot_path):
                img = XlImage(screenshot_path)
                img.width = 160
                img.height = 120
                ws.add_image(img, f"N{row_idx}")
                ws.row_dimensions[row_idx].height = 95
    except Exception as e:
        print(f"  [Warning] Could not embed screenshots: {e}")

    # -- Freeze panes --
    ws.freeze_panes = "A3"

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ Report saved: {output_path}")
